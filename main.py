import pandas as pd
import numpy as np
from scipy.optimize import linprog
import pulp
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import sys

class GrevesOptimizer:
    def __init__(self, excel_path):
        """Initialiser l'optimiseur à partir du fichier Excel"""
        self.excel_path = excel_path
        self.optimization_mode = None  # Track si Mode 1 ou Mode 2
        self.load_data()
        
    def load_data(self):
        """Charger les données du fichier Excel"""
        # Déterminer le nom de la feuille
        xls = pd.ExcelFile(self.excel_path)
        
        # Essayer de trouver la bonne feuille
        if "Données Grèves" in xls.sheet_names:
            sheet_name = "Données Grèves"
        elif "Planning" in xls.sheet_names:
            sheet_name = "Planning"
        else:
            # Utiliser la première feuille
            sheet_name = xls.sheet_names[0]
            print(f"⚠ Utilisation de la première feuille: {sheet_name}")
        
        # Lire sans header d'abord pour trouver les bonnes lignes
        df_raw = pd.read_excel(self.excel_path, sheet_name=sheet_name, header=None)
        
        # Trouver la ligne d'en-tête (contient "Prénom" et "Nom")
        header_row = None
        for idx in range(len(df_raw)):
            if df_raw.iloc[idx, 0] == 'Prénom' and df_raw.iloc[idx, 1] == 'Nom':
                header_row = idx
                break
        
        if header_row is None:
            raise ValueError("Impossible de trouver les colonnes 'Prénom' et 'Nom'")
        
        # Utiliser cette ligne comme header et ré-lire
        df = pd.read_excel(self.excel_path, sheet_name=sheet_name, header=header_row)
        
        # Nettoyer les données - retirer les lignes avec NaN ou "TOTAL"
        df = df.dropna(subset=['Prénom'], how='all')
        df = df[df['Prénom'].astype(str).notna()]
        df = df[df['Prénom'].astype(str).str.strip() != '']  # Retirer les lignes vides
        df = df[~df['Prénom'].astype(str).str.contains('TOTAL|TABLEAU|Période|INSTRUCTIONS|Coût|Unnamed', case=False, na=False)]
        
        # Extraire les enseignants (Prénom, Nom)
        self.teachers = df[['Prénom', 'Nom']].values
        
        # Formater les prénoms et noms en MAJUSCULES
        self.teachers = [[str(p).upper().strip(), str(n).upper().strip()] for p, n in self.teachers]
        # Retirer les lignes vides
        self.teachers = [pair for pair in self.teachers if pair[0] and pair[1] and pair[0] != 'NAN' and pair[1] != 'NAN']
        
        # Extraire les périodes de disponibilité (P1 à P10)
        periods = []
        for col in df.columns:
            if col and isinstance(col, str) and col.startswith('P'):
                try:
                    # Vérifier que ce qui suit 'P' est un nombre
                    period_num = int(col[1:])
                    periods.append(col)
                except:
                    pass
        
        self.periods = sorted(periods, key=lambda x: int(x[1:]))
        
        # Matrice de disponibilité (0, 1, ou 2)
        availability_data = df[self.periods].copy()
        # Convertir en nombres et remplacer les NaN par 0
        for col in self.periods:
            availability_data[col] = pd.to_numeric(availability_data[col], errors='coerce').fillna(0).astype(int)
        
        # VALIDATION : vérifier que les colonnes P1-P10 contiennent uniquement 0 ou 1
        invalid_values = []
        for i, row in enumerate(availability_data.values):
            for j, period in enumerate(self.periods):
                value = row[j]
                if value not in [0, 1]:
                    teacher_name = f"{self.teachers[i][0]} {self.teachers[i][1]}"
                    invalid_values.append(f"Ligne {i+2} ({teacher_name}), colonne {period}: valeur {value}")
        
        if invalid_values:
            error_msg = "❌ ERREUR DE VALIDATION : Les colonnes P1 à P10 doivent contenir uniquement 0 ou 1.\n\n"
            error_msg += "Valeurs invalides trouvées :\n"
            error_msg += "\n".join(invalid_values[:10])  # Limiter à 10 premières erreurs
            if len(invalid_values) > 10:
                error_msg += f"\n... et {len(invalid_values) - 10} autres erreurs"
            raise ValueError(error_msg)
        
        self.availability = availability_data.values
        
        # Extraire les besoins de grévistes par période
        self.required_strikers = {}
        
        # Méthode 1: Chercher les besoins dans les colonnes nommées (format original)
        if 'Nombre de grévistes nécessaire par période' in df.columns:
            # Chercher toutes les lignes où la colonne 'Nombre de grévistes' n'est pas vide
            needs_period_col = 'Nombre de grévistes nécessaire par période'
            # Chercher la colonne suivante qui contient les valeurs (normalement Unnamed: 15)
            value_col = None
            for col in df.columns:
                if col != needs_period_col and 'Unnamed' in str(col) and col.endswith('15'):
                    value_col = col
                    break
            
            # Si pas trouvée avec endswith, chercher juste après
            if value_col is None:
                col_idx = df.columns.tolist().index(needs_period_col)
                if col_idx + 1 < len(df.columns):
                    value_col = df.columns[col_idx + 1]
            
            if value_col:
                for idx, row in df.iterrows():
                    period = row[needs_period_col]
                    need = row[value_col]
                    if pd.notna(period) and pd.notna(need):
                        period_str = str(period).strip()
                        if period_str.startswith('P'):
                            try:
                                need_int = int(float(need))
                                self.required_strikers[period_str] = need_int
                            except:
                                pass
        
        
    def optimize(self, required_strikers=None):
        """Optimiser la sélection des grévistes"""
        # Marquer qu'on utilise le Mode 1
        self.optimization_mode = 1
        
        # Utiliser required_strikers passé en paramètre, sinon celui du fichier
        if required_strikers is not None:
            self.required_strikers = required_strikers
        elif not hasattr(self, 'required_strikers') or not self.required_strikers:
            raise ValueError("Aucun besoin spécifié. Fournir required_strikers en paramètre ou dans le fichier Excel.")
        
        
        num_teachers = len(self.teachers)
        num_periods = len(self.periods)
        
        # Vérification : pas de données
        if num_teachers == 0 or num_periods == 0:
            raise ValueError("Aucun enseignant ou période trouvé dans le fichier")
        
        # Créer le problème d'optimisation
        prob = pulp.LpProblem("Strike_Optimization", pulp.LpMinimize)
        
        # Variables de décision : strike[i][j] = 1 si l'enseignant i grève à la période j
        strike = [[pulp.LpVariable(f"strike_{i}_{j}", cat='Binary') 
                   for j in range(num_periods)] for i in range(num_teachers)]
        
        # Variables pour l'équité : min et max de périodes par enseignant
        max_periods_per_teacher = pulp.LpVariable("max_periods", lowBound=0, cat='Integer')
        min_periods_per_teacher = pulp.LpVariable("min_periods", lowBound=0, cat='Integer')
        
        # Objectif : minimiser l'écart entre max et min (équité parfaite),
        # puis le nombre total de grévistes-périodes (minimiser les heures perdues)
        spread = max_periods_per_teacher - min_periods_per_teacher
        obj = 10000 * spread + \
              100 * max_periods_per_teacher + \
              pulp.lpSum([strike[i][j] for i in range(num_teachers) for j in range(num_periods)])
        
        prob += obj
        
        # Contraintes de disponibilité
        for i in range(num_teachers):
            for j in range(num_periods):
                # On ne peut grever que si disponible (valeur 1 ou 2)
                if self.availability[i][j] == 0:
                    prob += strike[i][j] == 0
        
        # Contraintes de besoins : exactement le nombre de grévistes requis par période
        for j, period in enumerate(self.periods):
            if period in self.required_strikers:
                needed = self.required_strikers[period]
                # Nombre de grévistes == nombre requis (ni plus, ni moins)
                prob += pulp.lpSum([strike[i][j] for i in range(num_teachers)]) == needed
        
        # Contraintes d'équité : définir min et max de périodes par enseignant
        for i in range(num_teachers):
            periods_for_teacher = pulp.lpSum([strike[i][j] for j in range(num_periods)])
            prob += periods_for_teacher <= max_periods_per_teacher
            prob += periods_for_teacher >= min_periods_per_teacher
        
        # Résoudre
        print("\nRésolution du problème d'optimisation...")
        
        # Essayer différents solveurs dans l'ordre de préférence
        solvers_to_try = [
            ('CyLP', lambda: pulp.getSolver('CyLP', msg=0)),
            ('COIN_CMD', lambda: pulp.COIN_CMD(msg=0)),
            ('PULP_CBC_CMD', lambda: pulp.PULP_CBC_CMD(msg=0)),
        ]
        
        solved = False
        for solver_name, solver_func in solvers_to_try:
            try:
                solver = solver_func()
                if solver is not None and solver.available():
                    print(f"Utilisation du solveur: {solver_name}")
                    prob.solve(solver)
                    solved = True
                    break
            except Exception as e:
                print(f"⚠ {solver_name} non disponible: {e}")
                continue
        
        # Si aucun solveur PuLP ne fonctionne, utiliser scipy
        if not solved:
            print("⚠ Aucun solveur PuLP disponible, utilisation de scipy.optimize...")
            try:
                # Convertir le problème en format scipy
                solution_scipy = self._solve_with_scipy(num_teachers, num_periods)
                self.solution = solution_scipy
                print("✓ Solution trouvée avec scipy!")
                return solution_scipy
            except Exception as e:
                print(f"⚠ Erreur scipy: {e}")
                print("⚠ Utilisation du solveur glouton de secours...")
                try:
                    solution_greedy = self._solve_with_greedy(num_teachers, num_periods)
                    self.solution = solution_greedy
                    print("✓ Solution trouvée avec l'algorithme glouton!")
                    return solution_greedy
                except Exception as e2:
                    print(f"✗ Erreur algorithme glouton: {e2}")
                    raise RuntimeError("Aucun solveur d'optimisation disponible. Veuillez réinstaller l'application.")
        
        # Vérifier le statut de la solution
        if prob.status == 1 or prob.status == 0:
            print("✓ Solution optimale trouvée!")
        elif prob.status == -1:
            print("⚠ Le problème n'a pas de solution optimale (pas assez de disponibilités)")
            print("⚠ Utilisation du solveur glouton de secours...")
            solution_greedy = self._solve_with_greedy(num_teachers, num_periods)
            self.solution = solution_greedy
            print("✓ Solution approximative trouvée avec l'algorithme glouton!")
            return solution_greedy
        else:
            print(f"⚠ Problème de résolution (status: {prob.status})")
            print("⚠ Utilisation du solveur glouton de secours...")
            solution_greedy = self._solve_with_greedy(num_teachers, num_periods)
            self.solution = solution_greedy
            print("✓ Solution approximative trouvée avec l'algorithme glouton!")
            return solution_greedy
        
        # Extraire la solution
        solution = np.zeros_like(self.availability, dtype=int)
        for i in range(num_teachers):
            for j in range(num_periods):
                if strike[i][j].varValue == 1:
                    solution[i][j] = 2  # Marquer comme gréviste avec 2
        
        self.solution = solution
        return solution
    
    def _solve_with_scipy(self, num_teachers, num_periods):
        """Solveur de fallback utilisant scipy au lieu de PuLP"""
        from scipy.optimize import milp, LinearConstraint, Bounds
        
        # Créer les variables de décision (une variable binaire pour chaque enseignant/période)
        n_vars = num_teachers * num_periods
        
        # Fonction objectif : minimiser le nombre total de grévistes
        # On veut juste satisfaire les contraintes avec le minimum de personnes
        c = np.ones(n_vars)  # Tous les coefficients à 1
        
        # Contraintes de disponibilité : on ne peut grever que si disponible
        # Contraintes de besoins : nombre de grévistes >= besoin par période
        A_ub = []
        b_ub = []
        A_eq = []
        b_eq = []
        
        # Pour chaque période, contrainte de besoin minimum
        for j, period in enumerate(self.periods):
            if period in self.required_strikers:
                needed = self.required_strikers[period]
                # Somme des grévistes pour la période j >= needed
                # Transformé en: -somme <= -needed
                constraint = np.zeros(n_vars)
                for i in range(num_teachers):
                    if self.availability[i][j] > 0:  # Disponible
                        var_idx = i * num_periods + j
                        constraint[var_idx] = -1
                A_ub.append(constraint)
                b_ub.append(-needed)
        
        # Bounds: 0 ou 1 pour chaque variable, 0 si indisponible
        lb = np.zeros(n_vars)
        ub = np.ones(n_vars)
        for i in range(num_teachers):
            for j in range(num_periods):
                var_idx = i * num_periods + j
                if self.availability[i][j] == 0:
                    ub[var_idx] = 0  # Forcer à 0 si indisponible
        
        bounds = Bounds(lb, ub)
        
        # Variables entières (binaires)
        integrality = np.ones(n_vars, dtype=int)
        
        # Résoudre avec scipy
        if len(A_ub) > 0:
            constraints = LinearConstraint(np.array(A_ub), -np.inf, np.array(b_ub))
            result = milp(c=c, constraints=constraints, bounds=bounds, integrality=integrality)
        else:
            result = milp(c=c, bounds=bounds, integrality=integrality)
        
        if not result.success:
            raise ValueError("Scipy n'a pas trouvé de solution")
        
        # Convertir la solution en matrice
        solution = np.zeros_like(self.availability, dtype=int)
        for i in range(num_teachers):
            for j in range(num_periods):
                var_idx = i * num_periods + j
                if result.x[var_idx] > 0.5:  # Arrondi à 1
                    solution[i][j] = 2  # Marquer comme gréviste
        
        return solution
    
    def optimize_mode2(self, periods_per_teacher, closure_threshold=None, excluded_periods=None):
        """Mode 2 : Maximiser l'impact en respectant limite par enseignant
        
        # Marquer qu'on utilise le Mode 2
        self.optimization_mode = 2

        Args:
            periods_per_teacher: Nombre maximum de périodes grévées par enseignant
            closure_threshold: Seuil pour fermer une période (optionnel)
            excluded_periods: Liste des périodes à exclure

        Logique :
        - Maximise le nombre de périodes qui atteignent le seuil (si fourni)
        - Sinon, maximise le nombre total de grévistes
        - Équilibre la charge entre enseignants
        - IGNORE complètement TABLEAU 2 (required_strikers)
        """
        num_teachers = len(self.teachers)
        num_periods = len(self.periods)

        if num_teachers == 0 or num_periods == 0:
            raise ValueError("Aucun enseignant ou période trouvé dans le fichier")

        if excluded_periods is None:
            excluded_periods = []

        print(f"\n=== MODE 2 : Max {periods_per_teacher} périodes par enseignant ===")
        if closure_threshold:
            print(f"Seuil de fermeture : {closure_threshold} grévistes/période")
        if excluded_periods:
            print(f"Périodes exclues : {', '.join(excluded_periods)}")

        # Créer le problème d'optimisation
        prob = pulp.LpProblem("Strike_Mode2_Maximize", pulp.LpMaximize)

        # Variables de décision : strike[i][j] = 1 si enseignant i grève période j
        strike = [[pulp.LpVariable(f"strike_{i}_{j}", cat='Binary') 
                   for j in range(num_periods)] for i in range(num_teachers)]

        # Variable : nombre de grévistes par période
        strikers_per_period = [pulp.LpVariable(f"strikers_{j}", lowBound=0, cat='Integer')
                               for j in range(num_periods)]

        if closure_threshold:
            # Variable binaire : période j atteint le seuil (1) ou non (0)
            period_closed = [pulp.LpVariable(f"closed_{j}", cat='Binary')
                            for j in range(num_periods)]

            # Contrainte : period_closed[j] = 1 SSI strikers_per_period[j] >= threshold
            M = num_teachers  # Big M
            for j in range(num_periods):
                # Si closed = 1, alors strikers >= threshold
                prob += strikers_per_period[j] >= closure_threshold * period_closed[j]
                # Si closed = 1, alors strikers <= threshold (pas de surplus inutile)
                prob += strikers_per_period[j] <= closure_threshold + M * (1 - period_closed[j])
                # Si strikers >= threshold, alors closed doit être 1
                prob += period_closed[j] * M >= strikers_per_period[j] - closure_threshold + 1

            # OBJECTIF : Maximiser le nombre de périodes fermées
            # Objectif secondaire : minimiser l'écart-type de la charge (via max-min)
            max_load = pulp.LpVariable("max_load", lowBound=0, cat='Integer')

            for i in range(num_teachers):
                prob += pulp.lpSum([strike[i][j] for j in range(num_periods)]) <= max_load

            # Priorité 1 : maximiser périodes fermées (poids 1000)
            # Priorité 2 : minimiser charge max (poids 1)
            prob += 1000 * pulp.lpSum([period_closed[j] for j in range(num_periods)]) - max_load

        else:
            # Sans seuil : maximiser total en équilibrant la charge
            max_load = pulp.LpVariable("max_load", lowBound=0, cat='Integer')

            for i in range(num_teachers):
                prob += pulp.lpSum([strike[i][j] for j in range(num_periods)]) <= max_load

            # Maximiser total - charge max (pour équilibrer)
            total = pulp.lpSum([strikers_per_period[j] for j in range(num_periods)])
            prob += total - max_load

        # Contraintes de disponibilité
        for i in range(num_teachers):
            for j in range(num_periods):
                if self.availability[i][j] == 0:
                    prob += strike[i][j] == 0

        # Contrainte : périodes exclues (aucune grève)
        for period_name in excluded_periods:
            if period_name in self.periods:
                j = self.periods.index(period_name)
                prob += strikers_per_period[j] == 0

        # Contrainte : chaque enseignant AU MAXIMUM periods_per_teacher périodes
        for i in range(num_teachers):
            available_periods = sum(1 for j in range(num_periods) 
                                   if self.availability[i][j] > 0 and self.periods[j] not in excluded_periods)
            max_periods = min(periods_per_teacher, available_periods)
            prob += pulp.lpSum([strike[i][j] for j in range(num_periods)]) <= max_periods

        # Lier strikers_per_period au nombre de grévistes effectifs
        for j in range(num_periods):
            prob += strikers_per_period[j] == pulp.lpSum([strike[i][j] for i in range(num_teachers)])

        # Résoudre
        print("\nRésolution du problème d'optimisation...")

        solvers_to_try = [
            ('CyLP', lambda: pulp.getSolver('CyLP', msg=0)),
            ('COIN_CMD', lambda: pulp.COIN_CMD(msg=0, timeLimit=60)),
            ('PULP_CBC_CMD', lambda: pulp.PULP_CBC_CMD(msg=0, timeLimit=60)),
        ]

        solved = False
        for solver_name, solver_func in solvers_to_try:
            try:
                solver = solver_func()
                if solver is not None and solver.available():
                    print(f"Utilisation du solveur: {solver_name}")
                    prob.solve(solver)
                    solved = True
                    break
            except Exception as e:
                print(f"Échec du solveur {solver_name}: {e}")
                continue

        if not solved:
            print("Aucun solveur disponible, tentative avec le solveur par défaut...")
            prob.solve()

        # Vérifier le statut
        status = pulp.LpStatus[prob.status]
        print(f"\nStatut : {status}")

        if status != "Optimal":
            raise ValueError(f"Aucune solution optimale trouvée (statut : {status})")

        # Extraire la solution
        solution = np.zeros((num_teachers, num_periods), dtype=int)
        for i in range(num_teachers):
            for j in range(num_periods):
                if pulp.value(strike[i][j]) > 0.5:
                    solution[i][j] = 2
                else:
                    solution[i][j] = self.availability[i][j]

        self.solution = solution

        # Statistiques
        total = (solution == 2).sum()
        periods_with_strikes = sum(1 for j in range(num_periods) if (solution[:, j] == 2).sum() > 0)

        print(f"\n✓ Total grévistes-périodes : {total}")
        print(f"✓ Périodes avec grèves : {periods_with_strikes}/{num_periods}")

        if closure_threshold:
            periods_closed = sum(1 for j in range(num_periods) 
                               if (solution[:, j] == 2).sum() >= closure_threshold)
            print(f"✓ Périodes fermées (≥{closure_threshold} grévistes) : {periods_closed}")

        # Répartition par enseignant
        load_per_teacher = [(solution[i, :] == 2).sum() for i in range(num_teachers)]
        teachers_involved = sum(1 for load in load_per_teacher if load > 0)
        if teachers_involved > 0:
            avg_load = sum(load_per_teacher) / teachers_involved
            max_load_val = max(load_per_teacher)
            print(f"✓ Charge moyenne : {avg_load:.1f} périodes/enseignant")
            print(f"✓ Charge maximale : {max_load_val} périodes/enseignant")

        # Afficher répartition par période
        print("\nRépartition par période :")
        for j in range(num_periods):
            count = (solution[:, j] == 2).sum()
            period = self.periods[j]
            status = ""
            if period in excluded_periods:
                status = " [EXCLUE]"
            elif closure_threshold and count >= closure_threshold:
                status = " ✓ FERMÉE"
            elif closure_threshold and count > 0:
                status = f" ({count}/{closure_threshold})"
            print(f"  {period}: {count} grévistes{status}")

        return solution

    def _solve_with_greedy(self, num_teachers, num_periods):
        """Solveur de fallback ultime utilisant un algorithme glouton simple
        
        Cet algorithme fonctionne sans dépendances externes complexes.
        Il sélectionne greedily les enseignants qui minimisent la charge totale.
        """
        print("INFO: Utilisation d'un algorithme glouton optimisé")
        
        # Initialiser la solution à 0 (personne ne grève)
        solution = np.zeros_like(self.availability, dtype=int)
        
        # Pour chaque période, sélectionner les grévistes nécessaires
        for j, period in enumerate(self.periods):
            needed = self.required_strikers.get(period, 0)
            
            if needed == 0:
                continue
            
            # Créer une liste de candidats disponibles avec leur charge actuelle
            candidates = []
            for i in range(num_teachers):
                if self.availability[i][j] > 0:  # Disponible
                    # Compter combien de fois cet enseignant grève déjà
                    current_load = np.sum(solution[i, :] == 2)
                    candidates.append((i, current_load))
            
            if len(candidates) < needed:
                print(f"⚠ ATTENTION: Pas assez de candidats disponibles pour {period}")
                print(f"   Besoin: {needed}, Disponibles: {len(candidates)}")
                # Prendre tous les disponibles
                for i, _ in candidates:
                    solution[i][j] = 2
            else:
                # Trier les candidats par charge actuelle (les moins chargés en premier)
                # En cas d'égalité, ordre aléatoire pour équilibrer
                candidates.sort(key=lambda x: (x[1], np.random.random()))
                
                # Sélectionner les N premiers (avec charge minimale)
                for idx in range(needed):
                    teacher_idx, _ = candidates[idx]
                    solution[teacher_idx][j] = 2
        
        # Vérifier que toutes les contraintes sont satisfaites
        for j, period in enumerate(self.periods):
            needed = self.required_strikers.get(period, 0)
            actual = np.sum(solution[:, j] == 2)
            if actual < needed:
                print(f"⚠ AVERTISSEMENT: {period} n'a que {actual}/{needed} grévistes")
        
        print("INFO: Solution gloutonne générée avec succès")
        return solution
    
    def get_summary(self):
        """Obtenir un résumé de la solution"""
        summary = {}
        for j, period in enumerate(self.periods):
            strikers = []
            for i, (prenom, nom) in enumerate(self.teachers):
                if self.solution[i][j] == 2:
                    strikers.append(f"{prenom} {nom}")
            summary[period] = strikers
        return summary
    
    def save_to_excel(self, output_path):
        """Sauvegarder la solution optimisée en Excel"""
        # Préparer les données
        result_data = []
        for i, (prenom, nom) in enumerate(self.teachers):
            # S'assurer que les noms sont en majuscules
            prenom = str(prenom).upper().strip()
            nom = str(nom).upper().strip()
            row = {'Prénom': prenom, 'Nom': nom}
            total_strikes = 0
            for j, period in enumerate(self.periods):
                # Afficher "GREVE" pour les grévistes, "DISPONIBLE" pour disponibles non réquisitionnés
                if self.solution[i][j] == 2:
                    row[period] = "GREVE"
                    total_strikes += 1
                elif self.availability[i][j] == 1:
                    # Disponible mais pas réquisitionné
                    row[period] = "DISPONIBLE"
                else:
                    # Pas disponible - laisser vide
                    row[period] = ""
            row['Total Grèves'] = total_strikes
            result_data.append(row)
        
        result_df = pd.DataFrame(result_data)
        
        # Ajouter une ligne avec les totaux par période
        totals_row = {'Prénom': 'TOTAL', 'Nom': 'par période'}
        for j, period in enumerate(self.periods):
            total = sum(1 for i in range(len(self.teachers)) if self.solution[i][j] == 2)
            
            # Afficher différemment selon le mode
            if self.optimization_mode == 1:
                # Mode 1 : afficher total/besoin
                needed = self.required_strikers.get(period, 0)
                totals_row[period] = f"{total}/{int(needed)}" if needed > 0 else str(total)
            else:
                # Mode 2 : afficher seulement le total
                totals_row[period] = str(total)
        result_df = pd.concat([result_df, pd.DataFrame([totals_row])], ignore_index=True)
        
        # Créer un workbook avec mise en forme
        wb = Workbook()
        ws = wb.active
        ws.title = "Optimisation Grèves"
        
        # En-têtes
        headers = ['Prénom', 'Nom'] + self.periods + ["Total d'heures grévées par personne"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Données
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        white_font = Font(bold=True, color="FFFFFF")
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        for row_idx, row_data in enumerate(result_data, 2):
            ws.cell(row=row_idx, column=1, value=row_data['Prénom'])
            ws.cell(row=row_idx, column=2, value=row_data['Nom'])
            
            for col_idx, period in enumerate(self.periods, 3):
                value = row_data[period]
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if value == "GREVE":
                    cell.fill = red_fill
                    cell.font = white_font
                elif value == "DISPONIBLE":
                    cell.fill = green_fill
                    cell.font = white_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = border
            
            # Formule pour calculer automatiquement le total de grèves
            total_col_letter = chr(64 + len(headers))
            period_cols = [chr(64 + i) for i in range(3, 3 + len(self.periods))]
            formula = f'=COUNTIF({period_cols[0]}{row_idx}:{period_cols[-1]}{row_idx},"GREVE")'
            total_cell = ws.cell(row=row_idx, column=len(headers), value=formula)
            total_cell.alignment = Alignment(horizontal="center")
        
        # Ligne totaux
        last_row = len(result_data) + 2
        ws.cell(row=last_row, column=1, value='TOTAL')
        ws.cell(row=last_row, column=2, value='par période')
        ws.cell(row=last_row, column=1).font = Font(bold=True)
        ws.cell(row=last_row, column=2).font = Font(bold=True)
        
        for col_idx, period in enumerate(self.periods, 3):
            total = sum(1 for i in range(len(self.teachers)) if self.solution[i][col_idx-3] == 2)
            
            # Afficher différemment selon le mode
            if self.optimization_mode == 1:
                # Mode 1 : afficher total/besoin
                needed = self.required_strikers.get(period, 0)
                value = f"{total}/{int(needed)}" if needed > 0 else str(total)
            else:
                # Mode 2 : afficher seulement le total
                value = str(total)
            
            cell = ws.cell(row=last_row, column=col_idx, value=value)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        # Ajuster les largeurs de colonnes
        ws.column_dimensions['A'].width = 12  # Prénom
        ws.column_dimensions['B'].width = 12  # Nom
        # Largeur pour afficher "DISPONIBLE" confortablement
        for idx, period in enumerate(self.periods, 3):
            col_letter = chr(64 + idx) if idx < 27 else ''
            ws.column_dimensions[col_letter].width = 13
        # Largeur pour "Total d'heures grévées par personne"
        total_col = chr(64 + len(headers)) if len(headers) < 27 else 'N'
        ws.column_dimensions[total_col].width = 30
        
        # Ajouter mise en forme conditionnelle pour les couleurs dynamiques
        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.worksheet.datavalidation import DataValidation
        
        # Règle pour GREVE (rouge avec texte blanc)
        greve_rule = CellIsRule(
            operator='equal',
            formula=['"GREVE"'],
            fill=red_fill,
            font=white_font
        )
        
        # Règle pour DISPONIBLE (vert avec texte blanc)
        disponible_rule = CellIsRule(
            operator='equal',
            formula=['"DISPONIBLE"'],
            fill=green_fill,
            font=white_font
        )
        
        # Appliquer la mise en forme conditionnelle aux colonnes de périodes
        for col_idx in range(3, 3 + len(self.periods)):
            col_letter = chr(64 + col_idx)
            range_str = f'{col_letter}2:{col_letter}{len(result_data) + 1}'
            ws.conditional_formatting.add(range_str, greve_rule)
            ws.conditional_formatting.add(range_str, disponible_rule)
        
        # Ajouter validation pour les cellules de périodes (GREVE, DISPONIBLE ou vide)
        
        validation = DataValidation(
            type="list",
            formula1='"GREVE,DISPONIBLE,"',
            allow_blank=True
        )
        validation.error = 'Valeurs autorisées : GREVE, DISPONIBLE ou cellule vide'
        validation.errorTitle = 'Valeur invalide'
        validation.showErrorMessage = True
        validation.errorStyle = 'stop'
        
        # Appliquer la validation aux colonnes de périodes (lignes 2 à nombre d'enseignants + 1)
        for col_idx in range(3, 3 + len(self.periods)):
            col_letter = chr(64 + col_idx)
            validation.add(f'{col_letter}2:{col_letter}{len(result_data) + 1}')
        
        ws.add_data_validation(validation)
        
        # Créer le répertoire s'il n'existe pas
        import os
        output_dir = os.path.dirname(output_path)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir, exist_ok=True)
        
        wb.save(output_path)
        print(f"✓ Fichier sauvegardé: {output_path}")
    
    def interactive_adjustment(self):
        """Permettre l'ajustement interactif de la solution"""
        while True:
            print("\n" + "="*60)
            print("RÉSUMÉ DE LA SOLUTION")
            print("="*60)
            
            # Afficher la solution
            for i, (prenom, nom) in enumerate(self.teachers):
                prenom = str(prenom).upper().strip()
                nom = str(nom).upper().strip()
                strikes = [self.periods[j] for j in range(len(self.periods)) if self.solution[i][j] == 2]
                if strikes:
                    print(f"{prenom} {nom}: {', '.join(strikes)}")
            
            print("\n" + "-"*60)
            print("Options:")
            print("1. Accepter cette solution")
            print("2. Retirer une personne d'une période")
            print("3. Voir les statistiques")
            
            choice = input("\nChoisir une option (1-3): ").strip()
            
            if choice == "1":
                print("✓ Solution acceptée!")
                break
            elif choice == "2":
                self.remove_striker_from_period()
            elif choice == "3":
                self.show_statistics()
            else:
                print("Option invalide")
    
    def remove_striker_from_period(self):
        """Retirer une personne d'une période et proposer un remplaçant"""
        name = input("Nom de la personne à retirer (Format: Prénom Nom): ").strip().upper()
        period = input("Période à laquelle la retirer (P1-P10): ").strip()
        
        # Trouver la personne
        teacher_idx = None
        for i, (prenom, nom) in enumerate(self.teachers):
            prenom = str(prenom).upper().strip()
            nom = str(nom).upper().strip()
            if f"{prenom} {nom}" == name:
                teacher_idx = i
                break
        
        if teacher_idx is None:
            print("⚠ Enseignant non trouvé")
            return
        
        if period not in self.periods:
            print("⚠ Période invalide")
            return
        
        period_idx = self.periods.index(period)
        
        # Vérifier que cette personne était gréviste cette période
        if self.solution[teacher_idx][period_idx] != 2:
            print("⚠ Cette personne n'était pas gréviste cette période")
            return
        
        # Retirer
        self.solution[teacher_idx][period_idx] = self.availability[teacher_idx][period_idx]
        print(f"✓ {name} retiré de {period}")
        
        # Trouver un remplaçant
        self.find_replacement(period_idx)
    
    def find_replacement(self, period_idx, interactive=True, excluded_indices=None):
        """Trouver un remplaçant pour une période
        
        Args:
            period_idx: Index de la période
            interactive: Mode interactif (True) ou automatique (False)
            excluded_indices: Liste des indices d'enseignants à exclure (retraits manuels)
        """
        if excluded_indices is None:
            excluded_indices = []
            
        period = self.periods[period_idx]
        
        # Compter les grévistes actuels cette période
        current_strikers = sum(1 for i in range(len(self.teachers)) if self.solution[i][period_idx] == 2)
        needed = self.required_strikers.get(period, 0)
        
        if current_strikers >= needed:
            print(f"✓ Aucun remplaçant nécessaire ({current_strikers}/{int(needed)} grévistes)")
            return None
        
        # Chercher les candidats disponibles
        candidates = []
        for i, (prenom, nom) in enumerate(self.teachers):
            prenom = str(prenom).upper().strip()
            nom = str(nom).upper().strip()
            # Disponible si : pas encore gréviste cette période ET disponible (1 ou 2) ET pas exclu manuellement
            if (self.solution[i][period_idx] != 2 and 
                self.availability[i][period_idx] in [1, 2] and 
                i not in excluded_indices):
                # Compter combien de périodes cette personne gréverait
                future_strikes = sum(1 for j in range(len(self.periods)) if self.solution[i][j] == 2) + 1
                candidates.append((i, prenom, nom, future_strikes))
        
        if not candidates:
            print("⚠ Aucun candidat disponible pour remplacer!")
            return None
        
        # Trier par nombre de grèves (minimiser la charge)
        candidates.sort(key=lambda x: x[3])
        
        print(f"\nGrévistes proposés pour {period}:")
        for idx, (i, prenom, nom, future_strikes) in enumerate(candidates[:3], 1):
            print(f"  {idx}. {prenom} {nom} (grèverait {future_strikes} périodes au total)")
        
        if interactive:
            choice = input("Choisir un remplaçant (1-3) ou 'n' pour refuser: ").strip()
            
            if choice in ['1', '2', '3']:
                choice_idx = int(choice) - 1
                if choice_idx < len(candidates):
                    teacher_idx, prenom, nom, _ = candidates[choice_idx]
                    self.solution[teacher_idx][period_idx] = 2
                    print(f"✓ {prenom} {nom} sélectionné pour {period}")
                    return (teacher_idx, prenom, nom)
                else:
                    print("⚠ Choix invalide")
                    return None
            else:
                print("✓ Pas de remplaçant sélectionné")
                return None
        else:
            # Mode automatique : sélectionner le meilleur candidat
            teacher_idx, prenom, nom, _ = candidates[0]
            self.solution[teacher_idx][period_idx] = 2
            print(f"✓ {prenom} {nom} sélectionné automatiquement pour {period}")
            return (teacher_idx, prenom, nom)
    
    def show_statistics(self):
        """Afficher les statistiques de la solution"""
        print("\n" + "="*60)
        print("STATISTIQUES")
        print("="*60)
        
        # Statistiques par personne
        print("\nCharges par enseignant:")
        teacher_loads = []
        for i, (prenom, nom) in enumerate(self.teachers):
            prenom = str(prenom).upper().strip()
            nom = str(nom).upper().strip()
            strikes = sum(1 for j in range(len(self.periods)) if self.solution[i][j] == 2)
            teacher_loads.append((prenom, nom, strikes))
        
        teacher_loads.sort(key=lambda x: x[2], reverse=True)
        for prenom, nom, strikes in teacher_loads[:5]:
            print(f"  {prenom} {nom}: {strikes} périodes")
        
        # Statistiques par période
        print("\nGrévistes par période:")
        for j, period in enumerate(self.periods):
            strikers = sum(1 for i in range(len(self.teachers)) if self.solution[i][j] == 2)
            needed = self.required_strikers.get(period, 0)
            print(f"  {period}: {strikers}/{int(needed)} grévistes")
        
        # Total
        total_strikes = sum(sum(1 for j in range(len(self.periods)) if self.solution[i][j] == 2) 
                           for i in range(len(self.teachers)))
        active_periods = sum(1 for j in range(len(self.periods)) 
                            if sum(1 for i in range(len(self.teachers)) if self.solution[i][j] == 2) > 0)
        print(f"\nTotal de périodes grévées par le corps enseignant: {total_strikes}")
        print(f"Périodes avec grévistes: {active_periods}/{len(self.periods)}")

def main():
    print("Optimiseur de Grèves")
    print("="*60)
    
    # Proposer le choix du fichier
    print("\nFichiers disponibles:")
    print("1. planning_greve_reactif.xlsx (fichier original)")
    print("2. template_greve.xlsx (template à remplir)")
    print("3. Autre fichier (à saisir)")
    
    choice = input("\nChoisir le fichier à traiter (1, 2 ou 3): ").strip()
    
    if choice == "2":
        excel_path = "template_greve.xlsx"
    elif choice == "3":
        excel_path = input("Nom du fichier Excel: ").strip()
    else:
        excel_path = "planning_greve_reactif.xlsx"
    
    print(f"\n✓ Utilisation du fichier: {excel_path}")
    
    # Créer l'optimiseur
    optimizer = GrevesOptimizer(excel_path)
    
    # Optimiser
    optimizer.optimize()
    
    # Afficher les statistiques
    optimizer.show_statistics()
    
    # Sauvegarder d'abord
    output_path = input("\nNom du fichier de sortie (ex: resultat.xlsx): ").strip()
    if not output_path:
        output_path = "planning_greve_optimise.xlsx"
    optimizer.save_to_excel(output_path)
    
    # Proposer l'ajustement interactif après la sauvegarde
    print("\n" + "="*60)
    adjust = input("Voulez-vous ajuster la solution (remplacer des grévistes)? (o/n): ").strip().lower()
    if adjust in ['o', 'oui', 'y', 'yes']:
        optimizer.interactive_adjustment()
        # Sauvegarder à nouveau après les ajustements
        optimizer.save_to_excel(output_path)
        print(f"\n✓ Solution ajustée sauvegardée dans: {output_path}")
    
    print("\n✓ Optimisation terminée!")

if __name__ == "__main__":
    main()
