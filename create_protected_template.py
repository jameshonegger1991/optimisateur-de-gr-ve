"""
Script pour créer un template Excel protégé avec validation des données.
- Validation stricte : colonnes P1-P10 acceptent uniquement 0 ou 1
- Protection : seul le Tableau 1 (disponibilités) est modifiable
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, Protection
from openpyxl.worksheet.datavalidation import DataValidation
import random

def create_protected_template(filename="template_greve.xlsx", num_teachers=50, with_random_data=False):
    """
    Créer un template Excel protégé avec validation des données
    
    Args:
        filename: Nom du fichier à créer
        num_teachers: Nombre d'enseignants dans le template
        with_random_data: Si True, remplit avec des données aléatoires 0/1
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Données Grèves"
    
    # Styles
    header_fill = PatternFill(start_color="00D9FF", end_color="00D9FF", fill_type="solid")
    header_font = Font(bold=True, size=12, color="0A0E27")
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Titre principal
    ws.merge_cells('A1:L1')
    ws['A1'] = "OPTIMISATEUR DE GRÈVE - TABLEAU 1 : DISPONIBILITÉS"
    ws['A1'].font = Font(bold=True, size=14, color="00D9FF")
    ws['A1'].alignment = center_align
    
    # Instructions
    ws.merge_cells('A2:L2')
    ws['A2'] = "Remplissez uniquement les colonnes P1 à P10 avec 0 (pas disponible) ou 1 (disponible)"
    ws['A2'].font = Font(italic=True, size=10, color="FF0000")
    ws['A2'].alignment = center_align
    
    # En-têtes
    headers = ['Prénom', 'Nom'] + [f'P{i}' for i in range(1, 11)]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # Données exemples (lignes 4 à 4+num_teachers-1)
    # Liste de prénoms et noms français aléatoires
    prenoms = ["Marie", "Jean", "Sophie", "Pierre", "Julie", "Marc", "Laura", "Thomas", 
               "Céline", "Nicolas", "Emma", "Lucas", "Camille", "Alexandre", "Léa", 
               "Julien", "Sarah", "Mathieu", "Chloé", "David", "Manon", "Antoine",
               "Charlotte", "Vincent", "Lisa", "Maxime", "Océane", "Romain", "Eva",
               "Benjamin", "Alice", "François", "Clara", "Hugo", "Inès", "Arthur",
               "Jade", "Louis", "Anaïs", "Paul", "Lucie", "Simon", "Margot", "Gabriel",
               "Zoé", "Raphaël", "Louise", "Tom", "Lina", "Nathan"]
    
    noms = ["Martin", "Bernard", "Dubois", "Thomas", "Robert", "Richard", "Petit",
            "Durand", "Leroy", "Moreau", "Simon", "Laurent", "Lefebvre", "Michel",
            "Garcia", "David", "Bertrand", "Roux", "Vincent", "Fournier", "Morel",
            "Girard", "André", "Lefevre", "Mercier", "Dupont", "Lambert", "Bonnet",
            "François", "Martinez", "Legrand", "Garnier", "Faure", "Rousseau", "Blanc",
            "Guerin", "Muller", "Henry", "Roussel", "Nicolas", "Perrin", "Morin",
            "Mathieu", "Clement", "Gauthier", "Dumont", "Lopez", "Fontaine", "Chevalier", "Robin"]
    
    for row_idx in range(4, 4 + num_teachers):
        # Prénom et Nom aléatoires
        if with_random_data and row_idx - 4 < len(prenoms):
            ws.cell(row=row_idx, column=1).value = prenoms[row_idx - 4]
            ws.cell(row=row_idx, column=2).value = noms[row_idx - 4]
        else:
            ws.cell(row=row_idx, column=1).value = f"Prénom{row_idx-3}"
            ws.cell(row=row_idx, column=2).value = f"Nom{row_idx-3}"
        
        # P1 à P10 : valeurs aléatoires si demandé, sinon 0
        for col_idx in range(3, 13):
            cell = ws.cell(row=row_idx, column=col_idx)
            if with_random_data:
                # Générer 0 ou 1 aléatoirement (70% de chance d'avoir 1 pour plus de disponibilités)
                cell.value = random.choices([0, 1], weights=[30, 70])[0]
            else:
                cell.value = 0
            cell.alignment = center_align
            cell.border = border
    
    # VALIDATION DES DONNÉES : P1 à P10 doivent être 0 ou 1
    # Créer une validation de liste pour 0 ou 1
    dv = DataValidation(
        type="list",
        formula1='"0,1"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Valeur invalide",
        error="Seules les valeurs 0 ou 1 sont autorisées.\n0 = pas disponible\n1 = disponible"
    )
    
    # Appliquer la validation aux colonnes P1 à P10 (colonnes C à L)
    # Pour toutes les lignes de données (ligne 4 à 4+num_teachers-1)
    dv.add(f'C4:L{3 + num_teachers}')
    ws.add_data_validation(dv)
    
    # PROTECTION DE LA FEUILLE
    # 1. Verrouiller toutes les cellules par défaut
    for row in ws.iter_rows():
        for cell in row:
            cell.protection = Protection(locked=True)
    
    # 2. Déverrouiller UNIQUEMENT les cellules du Tableau 1 (P1 à P10)
    for row_idx in range(4, 4 + num_teachers):
        # Déverrouiller Prénom et Nom
        ws.cell(row=row_idx, column=1).protection = Protection(locked=False)
        ws.cell(row=row_idx, column=2).protection = Protection(locked=False)
        
        # Déverrouiller P1 à P10
        for col_idx in range(3, 13):
            ws.cell(row=row_idx, column=col_idx).protection = Protection(locked=False)
    
    # 3. Activer la protection de la feuille (sans mot de passe pour faciliter)
    ws.protection.sheet = True
    ws.protection.enable()
    
    # Ajuster les largeurs de colonnes
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
        ws.column_dimensions[col].width = 8
    
    # Note d'information en bas
    info_row = 4 + num_teachers + 2
    ws.merge_cells(f'A{info_row}:L{info_row}')
    ws[f'A{info_row}'] = "⚠️ PROTECTION ACTIVÉE : Seules les cellules du Tableau 1 peuvent être modifiées (Prénom, Nom, P1-P10)"
    ws[f'A{info_row}'].font = Font(italic=True, size=9, color="666666")
    ws[f'A{info_row}'].alignment = center_align
    
    # Sauvegarder
    wb.save(filename)
    print(f"✅ Template protégé créé : {filename}")
    print(f"   - {num_teachers} enseignants")
    print(f"   - Validation des données : P1-P10 acceptent uniquement 0 ou 1")
    print(f"   - Protection : seul le Tableau 1 est modifiable")


if __name__ == "__main__":
    # Créer le template vide (50 enseignants)
    create_protected_template("template_greve.xlsx", num_teachers=50, with_random_data=False)
    
    # Créer le template de test (50 enseignants avec données aléatoires)
    create_protected_template("template_greve_test_50.xlsx", num_teachers=50, with_random_data=True)
    
    print("\n🎯 Templates créés avec succès !")
